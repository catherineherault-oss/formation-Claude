#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export de TOUS les tableaux d'analyse dans un classeur Excel multi-feuilles.
Source : enquête 24407 (N=1025). Reproduit exactement les modèles du manuscrit
(code/03, code/06, code/07, code/10). Valeurs = résultats statistiques (non des
formules). Sortie : outputs/tableaux_analyses.xlsx
"""
import pandas as pd, numpy as np
from pathlib import Path
from scipy.cluster.hierarchy import linkage, fcluster
from scipy import stats
from sklearn.metrics import silhouette_score, calinski_harabasz_score
import statsmodels.formula.api as smf
import prince
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
A = pd.read_pickle(ROOT/"data/processed/analyse.pkl").reset_index(drop=True)
raw = pd.read_excel(ROOT/"data/raw/24407_Export.xlsx", sheet_name="Données_RepNat").reset_index(drop=True)
def num(c): return pd.to_numeric(raw[c], errors="coerce")
N = len(A)
CAN = {1:"Hyper/Super",2:"Hard discount",3:"Épiceries indép.",4:"Surgelés",5:"Bio spécialisé",
       6:"Marché",7:"Vente directe agri.",8:"Paniers interm.",9:"Artisans/comm.",
       10:"Coop./participatif",11:"Vrac/locaux",12:"Aide alimentaire",13:"Autre"}

def wilson(k,n,z=1.96):
    p=k/n; d=1+z**2/n; c=(p+z**2/(2*n))/d
    h=z*np.sqrt(p*(1-p)/n+z**2/(4*n**2))/d
    return 100*(c-h),100*(c+h)

# ---- Typologie commune (mêmes réglages que le manuscrit) -------------------
X=A[[f"cat_{i}" for i in range(1,12)]].astype(str)
mca=prince.MCA(n_components=5,random_state=42).fit(X)
MC=mca.row_coordinates(X).values
eig=list(mca.eigenvalues_); tot=sum(eig)
Z=linkage(MC,method="ward")
lab5=fcluster(Z,5,"maxclust")
order=pd.Series(lab5).value_counts().index.tolist()
A["classe"]=pd.Series(lab5,index=A.index).map({o:n for n,o in enumerate(order,1)})
NAME={1:"Multi-canal modérés",2:"Captifs de la grande distribution",3:"Multi-canal intensifs",
      4:"Explorateurs conventionnels",5:"Adeptes de la proximité"}

# ============================ Construction des tableaux =====================
TABLES=[]   # (nom_feuille, titre, [en-têtes], [lignes], {col:format})

# --- README ---
TABLES.append(("Lisez-moi","Tableaux d'analyse — vente directe (enquête 24407, N=1 025)",
    ["Élément","Détail"],
    [["Source","Enquête par questionnaire, échantillon représentatif national, N=1 025 ménages"],
     ["Généré par","code/13_export_tableaux.py (reproduit code/03, code/06, code/07, code/10)"],
     ["Nature des valeurs","Résultats statistiques agrégés (aucune donnée individuelle ; RGPD)"],
     ["Feuille clé demandée","« Régression VD inclusive » = tableau derrière la Figure 4 (odds-ratios)"],
     ["Définition vente directe","canal 7 (achat direct agriculteurs) ∪ Q22=1 (marché direct producteur)"],
     ["Note","% en points ; OR = odds-ratio ; IC 95 % de Wald (régressions) / Wilson (proportions)"]],
    {}))

# --- A1 Portefeuille de canaux ---
rows=[]
for i in range(1,14):
    rows.append([i,CAN[i],round(100*A[f"penet_{i}"].mean(),1),
                 round(100*A[f"reg_{i}"].mean(),1),round(A[f"part_{i}"].mean(),1)])
nbcan=A[[f'penet_{i}' for i in range(1,14)]].sum(axis=1).mean()
rows.append(["","Nb moyen de canaux fréquentés / ménage",round(nbcan,1),"",""])
TABLES.append(("A1_Portefeuille","A1 — Portefeuille de canaux (N=1 025)",
    ["#","Canal","Pénétration %","Réguliers ≥1×/mois %","Part budget moy. %"],rows,
    {3:"0.0",4:"0.0",5:"0.0"}))

# --- A2 Place de la vente directe ---
rows=[["VD via canal 7 (direct agriculteurs)",int(A.vd_canal7.sum()),round(100*A.vd_canal7.mean(),1)],
      ["VD via marché direct producteur (Q22=1)",int(A.vd_marche_direct.sum()),round(100*A.vd_marche_direct.mean(),1)],
      ["VD (union — définition retenue)",int(A.vente_directe.sum()),round(100*A.vente_directe.mean(),1)],
      ["dont apport propre de la porte marché",int(((A.vd_marche_direct==1)&(A.vd_canal7==0)).sum()),""],
      ["VD régulière (canal 7 ≥1×/mois)",int(A.vd_regulier.sum()),round(100*A.vd_regulier.mean(),1)]]
TABLES.append(("A2_Vente_directe","A2 — Place de la vente directe",
    ["Composante","n","% de l'échantillon"],rows,{3:"0.0"}))

# --- Sensibilité définition × dénominateur + IC Wilson ---
pen7=int(A.vd_canal7.sum()); union=int(A.vente_directe.sum()); reg7=int(A.vd_regulier.sum())
part_pop=A.part_vd_canal7.mean(); part_reg=A.loc[A.vd_regulier==1,"part_vd_canal7"].mean()
def wl(k): lo,hi=wilson(k,N); return f"[{lo:.1f}–{hi:.1f}]"
rows=[["Inclusive — union des 2 portes",union,round(100*union/N,1),wl(union),"—","—"],
      ["Pénétration canal 7 (toute fréquence)",pen7,round(100*pen7/N,1),wl(pen7),round(part_pop,1),"—"],
      ["Régulière (≥1×/mois)",reg7,round(100*reg7/N,1),wl(reg7),round(part_pop,1),round(part_reg,1)]]
TABLES.append(("Sensibilite","Sensibilité — définition × dénominateur budgétaire",
    ["Définition","n","% ménages","IC 95 % (Wilson)","Part budget /population %","Part budget /réguliers %"],
    rows,{3:"0.0",5:"0.0",6:"0.0"}))

# --- Typologie diagnostics ---
rows=[]
for k in (4,5,6):
    lab=fcluster(Z,k,"maxclust")
    eff=pd.Series(lab).value_counts().sort_values(ascending=False).tolist()
    rows.append([k,round(silhouette_score(MC,lab),3),int(round(calinski_harabasz_score(MC,lab))),str(eff)])
rows.append(["","","",""])
rows.append(["Axe ACM","Valeur propre","% inertie","% cumulé"])
cum=0
for i,e in enumerate(eig[:5],1):
    pct=100*e/tot; cum+=pct
    rows.append([i,round(e,4),round(pct,1),round(cum,1)])
TABLES.append(("Typologie_diagnostics","Diagnostics de partition (ACM + CAH Ward)",
    ["k (classes)","Silhouette moy.","Pseudo-F (Calinski-Harabasz)","Effectifs des classes"],rows,{}))

# --- Typologie 5 classes ---
pen=[f"penet_{i}" for i in range(1,14)]; regc=[f"reg_{i}" for i in range(1,14)]
rows=[]
for k in range(1,6):
    s=A[A.classe==k]
    rows.append([f"C{k} — {NAME[k]}",len(s),round(100*len(s)/N,1),round(s[pen].sum(axis=1).mean(),1),
                 round(s[regc].sum(axis=1).mean(),1),int(round(s.budget_eur.mean())),
                 round(100*s.vente_directe.mean(),0),round(100*s.vd_regulier.mean(),0)])
TABLES.append(("Typologie_5classes","Les cinq stratégies d'approvisionnement",
    ["Stratégie","n","% éch.","Canaux fréq. (moy.)","Canaux rég. (moy.)","Budget moy. €","% VD inclusive","% VD régulière"],
    rows,{3:"0.0",7:"0",8:"0"}))

# --- Usage régulier par canal × classe ---
rows=[]
for k in range(1,6):
    s=A[A.classe==k]
    rows.append([f"C{k}"]+[round(100*s[f"reg_{i}"].mean(),0) for i in range(1,12)])
TABLES.append(("Usage_reg_canal_classe","Usage régulier (≥1×/mois) par canal et par classe — % de la classe",
    ["Classe"]+[CAN[i] for i in range(1,12)],rows,{}))

# --- Profil socio des classes ---
def crosstab_rows(var):
    ct=pd.crosstab(A["classe"],A[var],normalize="index")*100
    p=stats.chi2_contingency(pd.crosstab(A["classe"],A[var]))[1]
    cols=list(ct.columns); out=[]
    for k in ct.index: out.append([f"C{k}"]+[round(ct.loc[k,c],0) for c in cols])
    return cols,out,p
socio=[]
for var,titre in [("csp","CSP"),("age","Âge"),("diplome","Diplôme")]:
    cols,out,p=crosstab_rows(var)
    socio.append(["— "+titre+f" × classe (%) — χ² p = {p:.4f} —"]+[""]*len(cols))
    socio.append(["Classe"]+[str(c) for c in cols])
    socio+=out
    socio.append([""]*(len(cols)+1))
# budget moyen par classe
socio.append(["— Budget moyen (€) par classe —",""])
for k in range(1,6):
    socio.append([f"C{k}",int(round(A[A.classe==k].budget_eur.mean()))])
TABLES.append(("Profil_socio_classes","Profil socio-économique des classes",
    ["Modalité / classe","valeurs →"],socio,{}))

# --- Régressions (VD inclusive = Figure 4, + VD régulière) ------------------
d=A.copy()
d["budget_z"]=(d.budget_eur-d.budget_eur.mean())/d.budget_eur.std()
d["age"]=pd.Categorical(d.age,["20-24","25-34","35-44","45-54","55-64","65+"])
d["csp"]=pd.Categorical(d.csp,["Employés/ouvriers","Indép./agri","Cadres/prof.lib","Prof. interm.","Retraités","Inactifs/autres"])
d["sexe"]=pd.Categorical(d.sexe,["Homme","Femme"])
d["diplome"]=pd.Categorical(d.diplome,["Infra-bac","Bac","Bac+2","Bac+3 et plus"])
formula="C(sexe)+C(age)+C(csp)+C(diplome)+budget_z"
if "tuu" in d.columns: d["tuu_z"]=(d.tuu-d.tuu.mean())/d.tuu.std(); formula+="+tuu_z"
def clean(idx): return idx.replace("C(","").replace(")","").replace("[T.","=").replace("]","")
def reg_table(dv):
    m=smf.logit(f"{dv} ~ {formula}",data=d).fit(disp=0)
    orv=np.exp(m.params); ci=np.exp(m.conf_int()); pv=m.pvalues
    rows=[]
    for i in m.params.index:
        if i=="Intercept": continue
        rows.append([clean(i),round(orv[i],2),round(ci.loc[i].iloc[0],2),round(ci.loc[i].iloc[1],2),
                     round(pv[i],3),"oui" if pv[i]<0.05 else "non"])
    rows.append(["","","","","",""])
    rows.append([f"Pseudo-R² (McFadden) = {m.prsquared:.3f} · N = {int(m.nobs)} · événements = {int(d[dv].sum())}","","","","",""])
    rows.append(["Réf. : Homme · 20-24 ans · Employés/ouvriers · Infra-bac","","","","",""])
    return rows
TABLES.append(("Regression_VD_inclusive","Déterminants de l'usage de la vente directe — VD INCLUSIVE (Figure 4)",
    ["Variable","Odds-ratio","IC 95 % bas","IC 95 % haut","p","Significatif (p<0,05)"],
    reg_table("vente_directe"),{2:"0.00",3:"0.00",4:"0.00",5:"0.000"}))
TABLES.append(("Regression_VD_reguliere","Déterminants — VD RÉGULIÈRE (≥1×/mois)",
    ["Variable","Odds-ratio","IC 95 % bas","IC 95 % haut","p","Significatif (p<0,05)"],
    reg_table("vd_regulier"),{2:"0.00",3:"0.00",4:"0.00",5:"0.000"}))

# --- Colinéarité ---
def cramers(a,b):
    ct=pd.crosstab(a,b); chi2=stats.chi2_contingency(ct)[0]; n=ct.values.sum(); r,k=ct.shape
    return np.sqrt((chi2/n)/(min(r-1,k-1)))
rho=stats.spearmanr(A.diplome.cat.codes.replace(-1,np.nan),A.budget_eur,nan_policy="omit").correlation
TABLES.append(("Colinearite","Colinéarité entre déterminants sociaux",
    ["Mesure","Valeur"],
    [["V de Cramér — diplôme × CSP",round(cramers(A.diplome,A.csp),2)],
     ["Corrélation de Spearman — diplôme × budget",round(rho,2)]],{2:"0.00"}))

# --- Récence par forme (mêmes réglages que code/07) ---
chans=[("Q100","Boulanger (conv.)"),("Q22","Marché (producteur)"),("Q92","Boucher (conv.)"),
       ("Q121","Primeur (conv.)"),("Q30","À la ferme"),("Q38","Magasin de producteurs"),
       ("Q61","Halle commerçante"),("Q70","Foire / salon"),("Q53","Panier en ligne")]
rows=[]
for q,lb in chans:
    s=num(q); base=s.notna().sum()
    rows.append([lb,int(base),round(100*(s==1).sum()/base,0),round(100*(s==2).sum()/base,0),
                 round(100*(s==3).sum()/base,0),round(100*(s>=4).sum()/base,0)])
TABLES.append(("Recence_par_forme","Récence d'achat par forme (% ligne)",
    ["Forme","Base n","Acheté le mois écoulé %","Achète, pas ce mois %","N'achète plus %","Jamais / ne connaît %"],
    rows,{}))

# --- AMAP (Q46) ---
q46=num("Q46"); base=q46.notna().sum(); oui=(q46==1).sum(); ab=(q46==2).sum(); ja=(q46==3).sum(); nc=(q46==4).sum()
taux=100*ab/(oui+ab); lo,hi=wilson(int(ab),int(oui+ab))
rows=[["Client actuel (1)",int(oui),round(100*oui/base,0)],
      ["A cessé (2)",int(ab),round(100*ab/base,0)],
      ["Jamais (3)",int(ja),round(100*ja/base,0)],
      ["Ne connaît pas l'AMAP (4)",int(nc),round(100*nc/base,0)],
      ["",""," "],
      [f"Base = {int(base)} acheteurs directs","",""],
      [f"Taux d'abandon = {ab}/({oui}+{ab}) = {taux:.0f}% (IC 95 % Wilson [{lo:.0f}–{hi:.0f}])","",""]]
TABLES.append(("AMAP_Q46","AMAP (Q46) — statut parmi les acheteurs directs",
    ["Statut","n","% base"],rows,{3:"0"}))

# --- Motivations marché (Q21) ---
MOT={1:"Découvrir nouveautés",2:"Flâner",3:"Grandes marques",4:"Produits locaux",5:"Soutien agriculteurs",
 6:"Rencontrer",7:"Diversité/choix",8:"Tout au même endroit",9:"Produits Bio",10:"Proximité",
 11:"Praticité",12:"Prix/promotions",13:"Qualité nutritionnelle",14:"Goût",15:"Circuit court",
 16:"Services",17:"Introuvables ailleurs",18:"Autre"}
top3=pd.concat([num(c) for c in ["Q21_1","Q21_2","Q21_3"]]); cnt=top3.value_counts()
mbase=raw["Q21_1"].notna().sum()
rows=[]
r=1
for item,c in cnt.head(10).items():
    if pd.isna(item) or item==0: continue
    rows.append([r,MOT.get(int(item),"?"),int(c),round(100*c/mbase,0)]); r+=1
TABLES.append(("Motivations_marche_Q21",f"Motivations d'achat sur les marchés (base : {int(mbase)}) — retiré du manuscrit",
    ["Rang","Motivation","Citations top 3","% base"],rows,{4:"0"}))

# --- Conversion ---
ratio=pen7/reg7; proj=part_pop*ratio
TABLES.append(("Conversion","Ordre de grandeur — gain de conversion (illustratif, borne haute)",
    ["Élément","Valeur"],
    [["Réguliers actuels (n / %)",f"{reg7} / {100*reg7/N:.1f}%"],
     ["Part nationale actuelle (canal 7)",f"{part_pop:.1f}%"],
     ["Intensité chez les réguliers",f"{part_reg:.1f}%"],
     ["Projection si essayeurs irréguliers → réguliers",f"≈ {proj:.1f}%"],
     ["Lecture","quasi-triplement ; borne haute, à valeur d'ordre de grandeur"]],{}))

# ============================ Écriture Excel ================================
wb=openpyxl.Workbook(); wb.remove(wb.active)
FT="Arial"
hdr_fill=PatternFill("solid",fgColor="1F4E78"); hdr_font=Font(name=FT,bold=True,color="FFFFFF",size=10)
title_font=Font(name=FT,bold=True,size=12,color="1F4E78")
cell_font=Font(name=FT,size=10)
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
for name,titre,headers,rows,fmts in TABLES:
    ws=wb.create_sheet(name[:31])
    ws["A1"]=titre; ws["A1"].font=title_font
    hr=3
    for j,h in enumerate(headers,1):
        c=ws.cell(hr,j,h); c.font=hdr_font; c.fill=hdr_fill
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    for r,row in enumerate(rows,hr+1):
        for j,val in enumerate(row,1):
            c=ws.cell(r,j,val); c.font=cell_font; c.border=border
            if j in fmts and isinstance(val,(int,float)): c.number_format=fmts[j]
            if j==1: c.alignment=Alignment(horizontal="left")
            else: c.alignment=Alignment(horizontal="center")
    # largeurs
    for j in range(1,len(headers)+1):
        maxlen=max([len(str(headers[j-1]))]+[len(str(row[j-1])) for row in rows if j-1<len(row)])
        ws.column_dimensions[get_column_letter(j)].width=min(max(maxlen+2,10),46)
    ws.freeze_panes="A4"
out=ROOT/"outputs/tableaux_analyses.xlsx"
wb.save(out)
print("Classeur écrit :",out)
print("Feuilles :",[n for n,*_ in TABLES])
